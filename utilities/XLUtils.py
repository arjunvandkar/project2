import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook[sheetName]
    row_count = sheet.max_row
    workbook.close()
    return row_count


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook[sheetName]
    col_count = sheet.max_column
    workbook.close()
    return col_count


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file, data_only=True)
    sheet = workbook[sheetName]
    value = sheet.cell(row=rownum, column=columnno).value
    workbook.close()
    return value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)
    workbook.close()


def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(
        start_color='60B212',
        end_color='60B212',
        fill_type='solid'
    )
    sheet.cell(row=rownum, column=columnno).fill = greenFill
    workbook.save(file)
    workbook.close()


def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(
        start_color='FF0000',
        end_color='FF0000',
        fill_type='solid'
    )
    sheet.cell(row=rownum, column=columnno).fill = redFill
    workbook.save(file)
    workbook.close()

